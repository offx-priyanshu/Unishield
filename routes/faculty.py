from flask import Blueprint, request, redirect, url_for, flash, render_template
from flask_login import login_required, current_user
from models.db import db
from models.outpass import Outpass
from models.log import Notification
from datetime import datetime, timedelta
from extensions import socketio

faculty_bp = Blueprint('faculty', __name__, url_prefix='/faculty')

@faculty_bp.route('/dashboard')
@login_required
def dashboard():
    if current_user.role not in ['hod', 'dean']:
        return redirect(url_for('auth.login'))
    
    # Calculate stats
    today = datetime.utcnow().replace(hour=0, minute=0, second=0)
    week_ago = datetime.utcnow() - timedelta(days=7)
    
    # All home-leave requests for stats
    all_home = Outpass.query.filter_by(pass_type='home')
    
    stats = {
        'approved_weekly': all_home.filter(Outpass.status != 'rejected', Outpass.status != 'pending', Outpass.created_at >= week_ago).count(),
        'rejected_today': all_home.filter(Outpass.status == 'rejected', Outpass.created_at >= today).count(),
        'emergency_count': all_home.filter(Outpass.priority == 'emergency', Outpass.status != 'approved', Outpass.status != 'rejected').count(),
        'parent_pending': all_home.filter(Outpass.parent_verified == False, Outpass.status != 'approved', Outpass.status != 'rejected').count()
    }
    
    # Show home-leave requests based on role and correct status
    if current_user.role == 'hod':
        pending_requests = Outpass.query.filter_by(pass_type='home', status='pending').all()
    else: # dean
        pending_requests = Outpass.query.filter_by(pass_type='home', status='hod_approved').all()
        
    return render_template('faculty/dashboard.html', requests=pending_requests, stats=stats)

@faculty_bp.route('/approve/<int:outpass_id>', methods=['POST'])
@login_required
def approve_request(outpass_id):
    outpass = Outpass.query.get_or_404(outpass_id)
    remark = request.form.get('remark', '')
    
    if current_user.role == 'hod':
        outpass.hod_id = current_user.id
        outpass.hod_signed_at = datetime.utcnow()
        outpass.hod_remarks = remark
        outpass.status = 'hod_approved'         # ← intermediate status
    elif current_user.role == 'dean':
        outpass.dean_id = current_user.id
        outpass.dean_signed_at = datetime.utcnow()
        outpass.dean_remarks = remark
        outpass.status = 'dean_approved'         # ← intermediate status
        
    # Add persistent notification for student
    new_notif = Notification(
        user_id=outpass.student_id,
        title=f"Leave Update: {current_user.role.upper()} Approved",
        message=f"Your outpass for {outpass.destination} has been approved by {current_user.name}. Proceeding to next stage.",
        type='success'
    )
    db.session.add(new_notif)
    db.session.commit()

    # Emit status update for real-time tracking
    socketio.emit('status_update', {
        'outpass_id': outpass.id,
        'student_id': outpass.student_id,
        'status': outpass.status,
        'approver': current_user.name,
        'role': current_user.role.upper(),
        'time': outpass.hod_signed_at.strftime('%H:%M') if current_user.role == 'hod' else outpass.dean_signed_at.strftime('%H:%M')
    }, namespace='/gate')

    flash(f"Request #{outpass_id} approved successfully.", "success")
    return redirect(url_for('faculty.dashboard'))

@faculty_bp.route('/reject/<int:outpass_id>', methods=['POST'])
@login_required
def reject_request(outpass_id):
    outpass = Outpass.query.get_or_404(outpass_id)
    remark = request.form.get('remark', 'Rejected by Faculty')
    
    outpass.status = 'rejected'
    if current_user.role == 'hod':
        outpass.hod_remarks = remark
    else:
        outpass.dean_remarks = remark
        
    db.session.commit()
    
    socketio.emit('status_update', {
        'outpass_id': outpass.id,
        'student_id': outpass.student_id,
        'status': 'rejected',
        'approver': current_user.name,
        'role': current_user.role.upper(),
        'reason': remark
    }, namespace='/gate')

    flash(f"Request #{outpass_id} has been rejected.", "danger")
    return redirect(url_for('faculty.dashboard'))

@faculty_bp.route('/verify_parent/<int:outpass_id>', methods=['POST'])
@login_required
def verify_parent(outpass_id):
    # This action is WARDEN-ONLY — HOD/Dean have NO verification authority
    # They can only enquire via phone. OTP verification is enforced at Warden level.
    flash('Parent verification is restricted to the Warden only. Please contact the Warden.', 'warning')
    return redirect(url_for('faculty.dashboard'))


# ─────────────────────────────────────────────
#  API & EXCEL EXPORTS (FACULTY)
# ─────────────────────────────────────────────
import io
from flask import make_response, jsonify

def _serialize_pass(p):
    s = p.student
    return {
        'id':              p.id,
        'student_name':    s.name        if s else 'N/A',
        'student_id':      s.student_id  if s else 'N/A',
        'department':      s.department  if s else 'N/A',
        'destination':     p.destination,
        'purpose':         p.purpose,
        'priority':        p.priority,
        'status':          p.status,
        'created_at':      p.created_at.strftime('%d %b %Y, %H:%M'),
        'parent_verified': p.parent_verified,
        'parent_phone':    s.parent_phone if s else 'N/A'
    }

def _build_excel_faculty(title_text, passes, filename_prefix):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Faculty Report'

    ws.merge_cells('A1:K1')
    tc = ws['A1']
    tc.value = title_text
    tc.font  = Font(name='Calibri', bold=True, size=13, color='FFFFFF')
    tc.fill  = PatternFill('solid', fgColor='0A0E12')
    tc.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    headers = ['#', 'Student Name', 'Roll No.', 'Dept', 'Destination', 'Purpose', 'Priority', 'Applied At', 'Parent Phone', 'Parent Verify', 'Status']
    hfill  = PatternFill('solid', fgColor='00D2FF')
    hfont  = Font(name='Calibri', bold=True, size=10, color='000000')
    border = Border(left=Side(style='thin', color='DDDDDD'), right=Side(style='thin', color='DDDDDD'), top=Side(style='thin', color='DDDDDD'), bottom=Side(style='thin', color='DDDDDD'))
    
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=ci, value=h)
        cell.font = hfont; cell.fill = hfill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    ws.row_dimensions[2].height = 20

    for ri, p in enumerate(passes, 1):
        row_data = [
            ri, p['student_name'], p['student_id'], p['department'],
            p['destination'], p['purpose'], p['priority'].upper(),
            p['created_at'], p['parent_phone'],
            'YES' if p['parent_verified'] else 'NO', p['status'].upper()
        ]
        er = ri + 2
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=er, column=ci, value=val)
            cell.border = border
            cell.font = Font(name='Calibri', size=9)
            if ci == 7 and val == 'EMERGENCY': cell.font = Font(bold=True, color='FF0000')

    for i, w in enumerate([5,22,14,12,20,30,12,18,15,12,18], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    resp = make_response(buf.getvalue())
    fname = f'{filename_prefix}_{datetime.utcnow().strftime("%Y%m%d")}.xlsx'
    resp.headers['Content-Disposition'] = f'attachment; filename={fname}'
    resp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    return resp

@faculty_bp.route('/api/approved-today')
@login_required
def api_approved_today():
    from datetime import timedelta
    week_ago = datetime.utcnow() - timedelta(days=7)
    passes = Outpass.query.filter(Outpass.status != 'rejected', Outpass.status != 'pending', Outpass.created_at >= week_ago).order_by(Outpass.created_at.desc()).all()
    return jsonify({'success':True, 'passes': [_serialize_pass(p) for p in passes]})

@faculty_bp.route('/export/approved-today')
@login_required
def export_approved_today():
    from datetime import timedelta
    week_ago = datetime.utcnow() - timedelta(days=7)
    passes = Outpass.query.filter(Outpass.status != 'rejected', Outpass.status != 'pending', Outpass.created_at >= week_ago).order_by(Outpass.created_at.desc()).all()
    return _build_excel_faculty('Approved Gate Passes (Last 7 Days)', [_serialize_pass(p) for p in passes], 'Faculty_Weekly_Report')

@faculty_bp.route('/api/emergency')
@login_required
def api_emergency():
    passes = Outpass.query.filter(Outpass.priority == 'emergency', Outpass.status != 'approved', Outpass.status != 'rejected').all()
    return jsonify({'success':True, 'passes': [_serialize_pass(p) for p in passes]})

@faculty_bp.route('/export/emergency')
@login_required
def export_emergency():
    passes = Outpass.query.filter(Outpass.priority == 'emergency', Outpass.status != 'approved', Outpass.status != 'rejected').all()
    return _build_excel_faculty('Active Emergency Requests', [_serialize_pass(p) for p in passes], 'Faculty_Emergency')

@faculty_bp.route('/api/parent-pending')
@login_required
def api_parent_pending():
    passes = Outpass.query.filter(Outpass.parent_verified == False, Outpass.status != 'approved', Outpass.status != 'rejected').all()
    return jsonify({'success':True, 'passes': [_serialize_pass(p) for p in passes]})

@faculty_bp.route('/export/parent-pending')
@login_required
def export_parent_pending():
    passes = Outpass.query.filter(Outpass.parent_verified == False, Outpass.status != 'approved', Outpass.status != 'rejected').all()
    return _build_excel_faculty('Awaiting Parent Verification', [_serialize_pass(p) for p in passes], 'Faculty_Parent_Pending')

