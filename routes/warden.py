from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, make_response
from flask_login import login_required, current_user
from models.db import db
from models.outpass import Outpass
from models.log import Notification
from datetime import datetime, timedelta
from extensions import socketio
import uuid, random, io

warden_bp = Blueprint('warden', __name__, url_prefix='/warden')

@warden_bp.route('/dashboard')
@login_required
def dashboard():
    if current_user.role != 'warden':
        return redirect(url_for('auth.login'))
    
    today = datetime.utcnow().replace(hour=0, minute=0, second=0)
    all_home = Outpass.query.filter_by(pass_type='home')
    
    stats = {
        'approved_today': all_home.filter(Outpass.status == 'approved', Outpass.warden_signed_at >= today).count(),
        'rejected_today': all_home.filter(Outpass.status == 'rejected', Outpass.created_at >= today).count(),
        'emergency_count': all_home.filter(Outpass.priority == 'emergency', Outpass.status == 'dean_approved').count(),
        'parent_pending': all_home.filter(Outpass.parent_verified == False, Outpass.status == 'dean_approved').count(),
        'total_passes_issued': Outpass.query.filter_by(status='approved').count()
    }
    
    # Warden sees home-leave requests that Dean has approved
    pending_requests = Outpass.query.filter_by(pass_type='home', status='dean_approved').all()
    return render_template('warden/dashboard.html', requests=pending_requests, stats=stats)

@warden_bp.route('/send-otp/<int:outpass_id>', methods=['POST'])
@login_required
def send_otp(outpass_id):
    if current_user.role != 'warden':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    outpass = Outpass.query.get_or_404(outpass_id)
    student = outpass.student
    
    if not student.parent_phone:
        return jsonify({'success': False, 'error': 'Parent phone number not registered for this student.'})
    
    # Generate 6-digit OTP
    otp = str(random.randint(100000, 999999))
    outpass.parent_otp = otp
    outpass.parent_verification_mode = 'otp'
    db.session.commit()
    
    # Send SMS to parent
    from services.sms_service import SMSService
    msg = (f"UniShield OTP: Your ward {student.name}'s home leave request has been approved "
           f"by HOD and Dean. Verification OTP: {otp}. Share with Warden to authorize gate pass.")
    success, result = SMSService.send_sms(student.parent_phone, msg, 'PARENT_OTP')
    
    if success:
        masked = student.parent_phone[:-4].replace(student.parent_phone[3:-4], 'X' * len(student.parent_phone[3:-4])) + student.parent_phone[-4:]
        return jsonify({'success': True, 'masked_phone': masked, 'message': f'OTP sent to {masked}'})
    else:
        # For demo/dev without SMS configured — still save OTP and return it
        masked = student.parent_phone[-4:].rjust(len(student.parent_phone), 'X')
        return jsonify({'success': True, 'masked_phone': masked, 'otp_debug': otp,
                        'message': f'SMS failed ({result}). Dev OTP: {otp}'})

@warden_bp.route('/verify-otp/<int:outpass_id>', methods=['POST'])
@login_required
def verify_otp(outpass_id):
    if current_user.role != 'warden':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    outpass = Outpass.query.get_or_404(outpass_id)
    entered_otp = request.json.get('otp', '').strip()
    
    if not outpass.parent_otp:
        return jsonify({'success': False, 'error': 'OTP not generated yet. Send OTP first.'})
    
    if entered_otp == outpass.parent_otp:
        outpass.parent_verified = True
        db.session.commit()
        return jsonify({'success': True, 'message': 'Parent verified successfully! QR generation unlocked.'})
    else:
        return jsonify({'success': False, 'error': 'Incorrect OTP. Please check and try again.'})

@warden_bp.route('/approve/<int:outpass_id>', methods=['POST'])
@login_required
def approve_request(outpass_id):
    if current_user.role != 'warden':
        return redirect(url_for('auth.login'))
    
    outpass = Outpass.query.get_or_404(outpass_id)
    
    if not outpass.parent_verified:
        flash("Parent OTP verification is mandatory before generating Gate Pass.", "danger")
        return redirect(url_for('warden.dashboard'))
        
    outpass.warden_id = current_user.id
    outpass.warden_signed_at = datetime.utcnow()
    outpass.parent_verified = True
    outpass.status = "approved"
    outpass.qr_token = str(uuid.uuid4())
    
    # Add persistent notification for student
    new_notif = Notification(
        user_id=outpass.student_id,
        title="Gate Pass Active",
        message=f"Final authorization granted by Warden {current_user.name}. Your QR Gate Pass is now active.",
        type='success'
    )
    db.session.add(new_notif)
    db.session.commit()

    # Emit status update for real-time tracking
    socketio.emit('status_update', {
        'outpass_id': outpass.id,
        'student_id': outpass.student_id,
        'status': outpass.status,
        'approver': current_user.name,
        'role': 'WARDEN',
        'time': outpass.warden_signed_at.strftime('%H:%M')
    }, namespace='/gate')

    flash(f"Final approval granted. Gate pass generated for Request #{outpass_id}.", "success")
    return redirect(url_for('warden.dashboard'))

@warden_bp.route('/reject/<int:outpass_id>', methods=['POST'])
@login_required
def reject_request(outpass_id):
    if current_user.role != 'warden':
        return redirect(url_for('auth.login'))
    
    outpass = Outpass.query.get_or_404(outpass_id)
    remark = request.form.get('remark', 'Rejected by Warden')
    
    outpass.status = 'rejected'
    outpass.dean_remarks = remark  # store in remarks
    
    new_notif = Notification(
        user_id=outpass.student_id,
        title="Gate Pass Denied",
        message=f"Warden {current_user.name} has rejected your leave request. Reason: {remark}",
        type='danger'
    )
    db.session.add(new_notif)
    db.session.commit()
    
    socketio.emit('status_update', {
        'outpass_id': outpass.id,
        'student_id': outpass.student_id,
        'status': 'rejected',
        'approver': current_user.name,
        'role': 'WARDEN',
        'reason': remark
    }, namespace='/gate')
    
    flash(f"Request #{outpass_id} rejected by Warden.", "danger")
    return redirect(url_for('warden.dashboard'))


@warden_bp.route('/weekly-passes')
@login_required
def weekly_passes():
    """Return all gate passes approved in the last 7 days as JSON."""
    if current_user.role != 'warden':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403

    week_ago = datetime.utcnow() - timedelta(days=7)
    passes = Outpass.query.filter(
        Outpass.status == 'approved',
        Outpass.warden_signed_at >= week_ago
    ).order_by(Outpass.warden_signed_at.desc()).all()

    data = []
    for p in passes:
        s = p.student
        data.append({
            'id':           p.id,
            'student_name': s.name if s else 'N/A',
            'student_id':   s.student_id if s else 'N/A',
            'department':   s.department if s else 'N/A',
            'destination':  p.destination,
            'purpose':      p.purpose,
            'priority':     p.priority,
            'start_date':   p.start_date.strftime('%d %b %Y') if p.start_date else '—',
            'end_date':     p.end_date.strftime('%d %b %Y') if p.end_date else '—',
            'approved_at':  p.warden_signed_at.strftime('%d %b %Y, %H:%M') if p.warden_signed_at else '—',
            'hod_name':     p.hod.name if p.hod else 'N/A',
            'dean_name':    p.dean.name if p.dean else 'N/A',
            'warden_name':  p.warden.name if p.warden else 'N/A',
            'parent_verified': p.parent_verified,
        })
    return jsonify({'success': True, 'passes': data, 'count': len(data)})


@warden_bp.route('/export-weekly-passes')
@login_required
def export_weekly_passes():
    """Export last 7-day approved gate passes as an Excel file."""
    if current_user.role != 'warden':
        flash('Unauthorized', 'danger')
        return redirect(url_for('auth.login'))

    week_ago = datetime.utcnow() - timedelta(days=7)
    passes = Outpass.query.filter(
        Outpass.status == 'approved',
        Outpass.warden_signed_at >= week_ago
    ).order_by(Outpass.warden_signed_at.desc()).all()

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Weekly Gate Passes'

        # --- Title row ---
        ws.merge_cells('A1:N1')
        title_cell = ws['A1']
        title_cell.value = f'UniShield — Weekly Gate Pass Report  ({(datetime.utcnow()-timedelta(days=7)).strftime("%d %b")} – {datetime.utcnow().strftime("%d %b %Y")})'
        title_cell.font = Font(name='Calibri', bold=True, size=14, color='FFFFFF')
        title_cell.fill = PatternFill('solid', fgColor='0D1117')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30

        # --- Header row ---
        headers = [
            '#', 'Student Name', 'Roll No.', 'Department', 'Destination',
            'Purpose', 'Priority', 'Start Date', 'End Date',
            'HOD Approved', 'Dean Approved', 'Warden Approved At',
            'Parent Verified', 'QR Token'
        ]
        header_fill  = PatternFill('solid', fgColor='1A7A4A')
        header_font  = Font(name='Calibri', bold=True, size=10, color='FFFFFF')
        thin_border  = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC'),
        )

        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
        ws.row_dimensions[2].height = 22

        # --- Data rows ---
        alt_fill = PatternFill('solid', fgColor='F2F9F5')
        for row_num, p in enumerate(passes, start=1):
            s = p.student
            row_data = [
                row_num,
                s.name if s else 'N/A',
                s.student_id if s else 'N/A',
                s.department if s else 'N/A',
                p.destination,
                p.purpose,
                p.priority.upper(),
                p.start_date.strftime('%d %b %Y') if p.start_date else '—',
                p.end_date.strftime('%d %b %Y') if p.end_date else '—',
                p.hod.name if p.hod else 'N/A',
                p.dean.name if p.dean else 'N/A',
                p.warden_signed_at.strftime('%d %b %Y %H:%M') if p.warden_signed_at else '—',
                'YES' if p.parent_verified else 'NO',
                p.qr_token or 'N/A',
            ]
            excel_row = row_num + 2
            fill = alt_fill if row_num % 2 == 0 else None
            for col_idx, val in enumerate(row_data, start=1):
                cell = ws.cell(row=excel_row, column=col_idx, value=val)
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center', wrap_text=True)
                cell.font = Font(name='Calibri', size=9)
                if fill:
                    cell.fill = fill
                # Colour Priority column
                if col_idx == 7:
                    if val == 'EMERGENCY':
                        cell.font = Font(name='Calibri', size=9, bold=True, color='CC0000')
                    elif val == 'MEDICAL':
                        cell.font = Font(name='Calibri', size=9, bold=True, color='0070C0')
                # Colour Parent Verified column
                if col_idx == 13:
                    cell.font = Font(name='Calibri', size=9, bold=True,
                                     color='1A7A4A' if val == 'YES' else 'CC0000')

        # --- Column widths ---
        col_widths = [5, 22, 14, 18, 22, 35, 11, 13, 13, 18, 18, 20, 14, 36]
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # --- Summary footer ---
        footer_row = len(passes) + 3
        ws.merge_cells(f'A{footer_row}:N{footer_row}')
        footer_cell = ws.cell(row=footer_row, column=1,
            value=f'Total Passes: {len(passes)}   |   Generated: {datetime.utcnow().strftime("%d %b %Y %H:%M")} UTC   |   UniShield Security System')
        footer_cell.font = Font(name='Calibri', italic=True, size=9, color='888888')
        footer_cell.alignment = Alignment(horizontal='right')

        # Save to buffer
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f'UniShield_GatePasses_{datetime.utcnow().strftime("%Y%m%d")}.xlsx'
        response = make_response(buf.getvalue())
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        return response

    except ImportError:
        flash('openpyxl library is required for Excel export. Run: pip install openpyxl', 'danger')
        return redirect(url_for('warden.dashboard'))


# ─────────────────────────────────────────────
#  Helper: serialize an outpass row
# ─────────────────────────────────────────────
def _serialize_pass(p):
    s = p.student
    return {
        'id':              p.id,
        'student_name':    s.name        if s else 'N/A',
        'student_id':      s.student_id  if s else 'N/A',
        'department':      s.department  if s else 'N/A',
        'destination':     p.destination,
        'purpose':         p.purpose,
        'priority':        p.priority,
        'status':          p.status,
        'start_date':      p.start_date.strftime('%d %b %Y')       if p.start_date      else '—',
        'end_date':        p.end_date.strftime('%d %b %Y')         if p.end_date        else '—',
        'approved_at':     p.warden_signed_at.strftime('%d %b %Y, %H:%M') if p.warden_signed_at else '—',
        'created_at':      p.created_at.strftime('%d %b %Y, %H:%M')  if p.created_at   else '—',
        'hod_name':        p.hod.name    if p.hod    else 'N/A',
        'dean_name':       p.dean.name   if p.dean   else 'N/A',
        'warden_name':     p.warden.name if p.warden else 'N/A',
        'parent_verified': p.parent_verified,
    }


def _build_excel(title_text, passes, filename_prefix):
    """Return a Flask Response with the formatted Excel file."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Gate Pass Report'

    ws.merge_cells('A1:L1')
    tc = ws['A1']
    tc.value = title_text
    tc.font  = Font(name='Calibri', bold=True, size=13, color='FFFFFF')
    tc.fill  = PatternFill('solid', fgColor='0D1117')
    tc.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    headers = ['#', 'Student Name', 'Roll No.', 'Department', 'Destination',
               'Purpose', 'Priority', 'Start Date', 'End Date',
               'Approved At', 'Parent Verified', 'Status']
    hfill  = PatternFill('solid', fgColor='1A7A4A')
    hfont  = Font(name='Calibri', bold=True, size=10, color='FFFFFF')
    border = Border(
        left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),  bottom=Side(style='thin', color='CCCCCC'),
    )
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=ci, value=h)
        cell.font = hfont; cell.fill = hfill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    ws.row_dimensions[2].height = 20

    alt = PatternFill('solid', fgColor='F2F9F5')
    for ri, p in enumerate(passes, 1):
        row_data = [
            ri, p['student_name'], p['student_id'], p['department'],
            p['destination'], p['purpose'], p['priority'].upper(),
            p['start_date'], p['end_date'], p['approved_at'] or p['created_at'],
            'YES' if p['parent_verified'] else 'NO', p['status'].replace('_',' ').upper(),
        ]
        er = ri + 2
        fill = alt if ri % 2 == 0 else None
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=er, column=ci, value=val)
            cell.border = border
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            cell.font = Font(name='Calibri', size=9)
            if fill: cell.fill = fill
            if ci == 7 and val == 'EMERGENCY':
                cell.font = Font(name='Calibri', size=9, bold=True, color='CC0000')
            if ci == 7 and val == 'MEDICAL':
                cell.font = Font(name='Calibri', size=9, bold=True, color='0070C0')
            if ci == 11:
                cell.font = Font(name='Calibri', size=9, bold=True,
                                 color='1A7A4A' if val == 'YES' else 'CC0000')

    for i, w in enumerate([5,22,14,18,22,35,11,13,13,20,13,18], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    fr = len(passes) + 3
    ws.merge_cells(f'A{fr}:L{fr}')
    fc = ws.cell(row=fr, column=1,
        value=f'Total: {len(passes)}  |  Generated: {datetime.utcnow().strftime("%d %b %Y %H:%M")} UTC  |  UniShield')
    fc.font = Font(name='Calibri', italic=True, size=9, color='888888')
    fc.alignment = Alignment(horizontal='right')

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    resp = make_response(buf.getvalue())
    fname = f'{filename_prefix}_{datetime.utcnow().strftime("%Y%m%d")}.xlsx'
    resp.headers['Content-Disposition'] = f'attachment; filename={fname}'
    resp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    return resp


# ─────────────────────────────────────────────
#  Awaiting Seal — JSON + Excel
# ─────────────────────────────────────────────
@warden_bp.route('/awaiting-passes')
@login_required
def awaiting_passes():
    if current_user.role != 'warden':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    passes = Outpass.query.filter_by(pass_type='home', status='dean_approved').all()
    return jsonify({'success': True, 'passes': [_serialize_pass(p) for p in passes], 'count': len(passes)})


@warden_bp.route('/export-awaiting-passes')
@login_required
def export_awaiting_passes():
    if current_user.role != 'warden':
        return redirect(url_for('auth.login'))
    passes = Outpass.query.filter_by(pass_type='home', status='dean_approved').all()
    try:
        return _build_excel('UniShield — Pending Warden Seal Requests', [_serialize_pass(p) for p in passes], 'UniShield_Awaiting')
    except ImportError:
        flash('openpyxl required. Run: pip install openpyxl', 'danger')
        return redirect(url_for('warden.dashboard'))


# ─────────────────────────────────────────────
#  Today's Passes — JSON + Excel
# ─────────────────────────────────────────────
@warden_bp.route('/today-passes')
@login_required
def today_passes():
    if current_user.role != 'warden':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    today = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
    passes = Outpass.query.filter(
        Outpass.status == 'approved',
        Outpass.warden_signed_at >= today
    ).order_by(Outpass.warden_signed_at.desc()).all()
    return jsonify({'success': True, 'passes': [_serialize_pass(p) for p in passes], 'count': len(passes)})


@warden_bp.route('/export-today-passes')
@login_required
def export_today_passes():
    if current_user.role != 'warden':
        return redirect(url_for('auth.login'))
    today = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
    passes = Outpass.query.filter(
        Outpass.status == 'approved',
        Outpass.warden_signed_at >= today
    ).order_by(Outpass.warden_signed_at.desc()).all()
    title = f"UniShield — Today's Gate Passes ({datetime.utcnow().strftime('%d %b %Y')})"
    try:
        return _build_excel(title, [_serialize_pass(p) for p in passes], 'UniShield_Today')
    except ImportError:
        flash('openpyxl required. Run: pip install openpyxl', 'danger')
        return redirect(url_for('warden.dashboard'))


# ─────────────────────────────────────────────
#  Emergency Passes — JSON + Excel
# ─────────────────────────────────────────────
@warden_bp.route('/emergency-passes')
@login_required
def emergency_passes():
    if current_user.role != 'warden':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    passes = Outpass.query.filter(
        Outpass.priority == 'emergency',
        Outpass.status.in_(['pending', 'hod_approved', 'dean_approved'])
    ).order_by(Outpass.created_at.desc()).all()
    return jsonify({'success': True, 'passes': [_serialize_pass(p) for p in passes], 'count': len(passes)})


@warden_bp.route('/export-emergency-passes')
@login_required
def export_emergency_passes():
    if current_user.role != 'warden':
        return redirect(url_for('auth.login'))
    passes = Outpass.query.filter(
        Outpass.priority == 'emergency',
        Outpass.status.in_(['pending', 'hod_approved', 'dean_approved'])
    ).order_by(Outpass.created_at.desc()).all()
    try:
        return _build_excel('UniShield — Emergency Requests', [_serialize_pass(p) for p in passes], 'UniShield_Emergency')
    except ImportError:
        flash('openpyxl required. Run: pip install openpyxl', 'danger')
        return redirect(url_for('warden.dashboard'))
